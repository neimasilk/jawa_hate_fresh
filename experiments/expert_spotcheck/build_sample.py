"""Build the native-expert spot-check (fallback-ladder #2).

Selects a stratified 100-item sample from the CLEAN (PII-scrubbed) consensus file,
stratified on (LLM consensus_hate x source orig_label), shuffled and BLIND (no LLM
or source label shown), and writes an Excel form for the expert (Bapak) to fill.
A hidden key (gitignored) lets score_spotcheck.py compute agreement afterwards.

Strata (the informative ones for the kappa=0.19 question):
  C  LLM=hate, source=neutral  -> ALL 24  (the slurs humans missed; strongest claim)
  D  LLM=non-hate, source=hate -> 40       (riskiest: is the LLM wrongly dropping hate?)
  A  LLM=hate, source=hate     -> 20       (control / sanity)
  B  LLM=non-hate, source=neut -> 16       (control / sanity)

Run: python experiments/expert_spotcheck/build_sample.py
"""
from __future__ import annotations

import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
SRC = ROOT / "data/labeled/bulk_v2_consensus_clean.jsonl"
OUT = Path(__file__).parent
SEED = 20260623
N = {"C": 24, "D": 40, "A": 20, "B": 16}

INSTRUKSI = [
    ("PETUNJUK SPOT-CHECK PAKAR — Ujaran Kebencian Jawa", True),
    ("", False),
    ("Tujuan: Bapak (penutur asli Jawa) menilai 100 teks secara MANDIRI untuk", False),
    ("menjadi jangkar validitas. Ini BUKAN anotasi penuh — cukup ~1 jam.", False),
    ("", False),
    ("CARA: buka sheet 'Penilaian'. Untuk tiap teks, isi kolom KUNING:", False),
    ("  - 'Ujaran Kebencian?'  ->  ketik 1 (ya) atau 0 (tidak). WAJIB.", False),
    ("  - 'Target grup' & 'Catatan'  ->  opsional (boleh kosong).", False),
    ("PENTING: nilai MURNI penilaian Bapak. Tidak ada label LLM/sumber ditampilkan,", False),
    ("supaya tidak terpengaruh. Jangan cari 'jawaban benar' — naluri pakar = data.", False),
    ("", False),
    ("DEFINISI (ringkas dari codebook):", True),
    ("Ujaran kebencian (isi 1) = menyerang/merendahkan/mendehumanisasi/menghasut", False),
    ("terhadap orang atau kelompok KARENA identitas kelompoknya (suku, agama/", False),
    ("kepercayaan, gender/orientasi seksual, kelas sosial, atau kelompok politik).", False),
    ("", False),
    ("BUKAN hate (isi 0) — walau sangat kasar:", True),
    ("  - Umpatan/pisuhan (asu, jancuk, goblok) tanpa dimensi kelompok; banter teman.", False),
    ("  - Kritik kasar atas KINERJA pejabat/institusi (bukan identitasnya).", False),
    ("  - Keluhan/makian ke situasi, benda, atau diri sendiri.", False),
    ("", False),
    ("TAPI individu BISA hate kalau diserang pakai slur identitas:", True),
    ("  - perempuan: lonte, murahan, 'rahim anget'  -> hate", False),
    ("  - agama: kapir/kafir, bani micin  -> hate", False),
    ("  - orientasi: banci, maho  -> hate", False),
    ("", False),
    ("TES CEPAT kalau ragu: 'Apakah ini merendahkan KELOMPOK identitas,", True),
    ("atau seseorang KARENA identitas kelompoknya?' Kalau TIDAK -> isi 0.", False),
    ("", False),
    ("Setelah selesai, simpan file ini dan kirim balik. Saya hitung otomatis", False),
    ("kesepakatan Bapak vs LLM dan vs label sumber, lalu masukkan ke paper.", False),
]


def main():
    rows = [json.loads(l) for l in SRC.open(encoding="utf-8")]
    buckets = {"A": [], "B": [], "C": [], "D": []}
    for r in rows:
        llm = r.get("consensus_hate")
        src = r.get("orig_label") == "hate"
        if llm and src:
            buckets["A"].append(r)
        elif (not llm) and (not src):
            buckets["B"].append(r)
        elif llm and (not src):
            buckets["C"].append(r)
        elif (not llm) and src:
            buckets["D"].append(r)

    rng = random.Random(SEED)
    sample = []
    for strat, k in N.items():
        pool = buckets[strat][:]
        rng.shuffle(pool)
        sample += [(strat, r) for r in pool[:k]]
    rng.shuffle(sample)  # blind: mix strata

    # ---- hidden key (gitignored) ----
    key_lines = ["no,id,stratum,llm_consensus_hate,orig_label"]
    for i, (strat, r) in enumerate(sample, 1):
        key_lines.append(f"{i},{r['source_id']},{strat},{int(bool(r.get('consensus_hate')))},{r.get('orig_label')}")
    (OUT / "_key.csv").write_text("\n".join(key_lines) + "\n", encoding="utf-8")

    # ---- expert form (xlsx) ----
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Petunjuk"
    ws0.column_dimensions["A"].width = 95
    for i, (txt, bold) in enumerate(INSTRUKSI, 1):
        c = ws0.cell(row=i, column=1, value=txt)
        c.font = Font(bold=bold, size=12 if (bold and i == 1) else 11)
        c.alignment = Alignment(wrap_text=True)

    ws = wb.create_sheet("Penilaian")
    headers = ["No", "ID", "Teks",
               "Ujaran Kebencian? (1=ya / 0=tidak)", "Target grup (opsional)", "Catatan (opsional)"]
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="D9E1F2")
    yfill = PatternFill("solid", fgColor="FFF2CC")
    for col in range(1, 7):
        cc = ws.cell(row=1, column=col)
        cc.font = Font(bold=True)
        cc.fill = yfill if col in (4, 5, 6) else hfill
        cc.alignment = Alignment(wrap_text=True, vertical="center")
    widths = [5, 22, 70, 16, 18, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    for i, (strat, r) in enumerate(sample, 1):
        ws.append([i, r["source_id"], r["text"], "", "", ""])
        ws.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (4, 5, 6):
            ws.cell(row=i + 1, column=col).fill = yfill

    dv = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"D2:D{len(sample) + 1}")
    ws.freeze_panes = "A2"

    out_xlsx = OUT / "SPOTCHECK_FORM.xlsx"
    wb.save(out_xlsx)
    print(f"Wrote {out_xlsx.name}: {len(sample)} items "
          f"(A={N['A']} B={N['B']} C={N['C']} D={N['D']})")
    print(f"Hidden key: _key.csv (gitignored)")


if __name__ == "__main__":
    main()
