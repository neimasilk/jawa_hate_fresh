"""Score the filled expert spot-check against LLM consensus and source labels.

Run AFTER Bapak fills SPOTCHECK_FORM.xlsx (column D = 1/0).
  python experiments/expert_spotcheck/score_spotcheck.py

Computes expert-vs-LLM and expert-vs-source agreement + Cohen kappa (overall and
per stratum), and the decisive question: on the cases where LLM and source DISAGREE,
whom does the native expert side with? Writes spotcheck_result.md.
"""
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

OUT = Path(__file__).parent
FORM = OUT / "SPOTCHECK_FORM.xlsx"
KEY = OUT / "_key.csv"


def cohen_kappa(pairs):
    n = len(pairs)
    if n == 0:
        return float("nan")
    pa = sum(1 for a, b in pairs if a == b) / n
    pay = sum(1 for a, _ in pairs if a) / n
    pby = sum(1 for _, b in pairs if b) / n
    pe = pay * pby + (1 - pay) * (1 - pby)
    return (pa - pe) / (1 - pe) if pe != 1 else float("nan")


def main():
    key = {}
    with KEY.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key[row["id"]] = row

    wb = load_workbook(FORM)
    ws = wb["Penilaian"]
    expert = {}
    missing = []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None:
            continue
        val = ws.cell(r, 4).value
        if val in (None, ""):
            missing.append(str(no))
            continue
        expert[str(no)] = 1 if str(val).strip() in ("1", "1.0", "ya", "Ya") else 0

    out = []
    P = lambda *a: out.append(" ".join(str(x) for x in a))
    P("# Expert Spot-Check Result\n")
    if missing:
        P(f"> WARNING: {len(missing)} items unlabeled (no {','.join(missing[:20])}...). Fill them.\n")

    rows = []
    for no, e in expert.items():
        k = key.get(no)
        if not k:
            continue
        rows.append(dict(no=no, expert=e, llm=int(k["llm_consensus_hate"]),
                         src=1 if k["orig_label"] == "hate" else 0, strat=k["stratum"]))
    P(f"Scored {len(rows)} items.\n")

    # overall
    e_llm = [(r["expert"], r["llm"]) for r in rows]
    e_src = [(r["expert"], r["src"]) for r in rows]
    P("## Overall agreement\n")
    P(f"- Expert vs LLM consensus: {sum(1 for a,b in e_llm if a==b)}/{len(e_llm)} "
      f"= {sum(1 for a,b in e_llm if a==b)/len(e_llm):.1%}, Cohen kappa {cohen_kappa(e_llm):.3f}")
    P(f"- Expert vs source labels: {sum(1 for a,b in e_src if a==b)}/{len(e_src)} "
      f"= {sum(1 for a,b in e_src if a==b)/len(e_src):.1%}, Cohen kappa {cohen_kappa(e_src):.3f}\n")

    # the decisive question: divergent cases
    div = [r for r in rows if r["llm"] != r["src"]]
    side_llm = sum(1 for r in div if r["expert"] == r["llm"])
    P("## DECISIVE: on cases where LLM and source DISAGREE\n")
    P(f"- {len(div)} divergent items; expert sided with **LLM {side_llm}** vs source {len(div)-side_llm} "
      f"= **{side_llm/len(div):.0%} for the LLM**\n" if div else "- (no divergent items)\n")

    # per stratum
    P("## Per stratum\n")
    P("| stratum | meaning | n | expert agrees LLM | expert agrees source |")
    P("|---|---|---|---|---|")
    meanings = {"A": "LLM-hate & src-hate", "B": "LLM-non & src-neutral",
                "C": "LLM-hate & src-neutral (slurs humans missed)",
                "D": "LLM-non & src-hate (LLM dropped human-hate)"}
    for s in ["A", "B", "C", "D"]:
        sr = [r for r in rows if r["strat"] == s]
        if not sr:
            continue
        al = sum(1 for r in sr if r["expert"] == r["llm"])
        asrc = sum(1 for r in sr if r["expert"] == r["src"])
        P(f"| {s} | {meanings[s]} | {len(sr)} | {al}/{len(sr)} ({al/len(sr):.0%}) | {asrc}/{len(sr)} ({asrc/len(sr):.0%}) |")

    P("\n## Interpretation guide")
    P("- Stratum C: expert agreeing with LLM (=hate) CONFIRMS the pipeline catches real slurs humans missed.")
    P("- Stratum D: expert agreeing with LLM (=non-hate) CONFIRMS the narrower definition; "
      "expert saying hate => real false negatives to fix.")
    P("- High expert-vs-LLM kappa + majority siding with LLM on divergent cases => "
      "the kappa-0.19 vs source becomes evidence the source labels are noisier, not that the LLMs are wrong.")

    report = "\n".join(out)
    (OUT / "spotcheck_result.md").write_text(report, encoding="utf-8")
    print(report)


if __name__ == "__main__":
    main()
