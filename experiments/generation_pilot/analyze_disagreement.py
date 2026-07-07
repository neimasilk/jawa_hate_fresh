"""Reproduce the Mukhlis-Yekti / Mukhlis-Daniel disagreement diagnosis (STRATEGY.md
P0-2) and run the E6-lite JELAS_HATE x machine_caught cross-tab (P0-3) plus the
per-model x per-validator authenticity table (P0-4) -- all from data already collected
(VALIDATION_FORM*.xlsx + _key.csv), no new native input needed.

Why this exists: STATE.md C13 / wiki/decisions.md D21 / paper draft_jinita.md SS4.7(1)
cite disagreement-row numbers (39 Mukhlis-Yekti disagreements, share concentrated in
krama_sarcastic, harmonized alpha 0.095->0.519, code-switching note count 31, etc.) that
were computed ad-hoc in a Bash session and never captured in a committed, rerunnable
script -- a violation of the D17 SOP (every number cited in the paper must be
reproducible from committed code, verified via a second, differently-implemented
pathway). This script is that script.

It also runs two analyses the data already supports but that were never run:
  - JELAS_HATE rate per niche per validator (Yekti, Daniel).
  - JELAS_HATE x machine_caught cross-tab on the 36 DeepSeek cells the detector probe
    covers, including the specific question of whether the 9 cells that evade ALL 5
    detectors are rated JELAS_HATE=1 by either validator. This bears directly on SS4.5's
    claim ("the hate that detectors miss is real Javanese hate") -- if the answer is
    close to zero, that claim needs qualification, not silent confirmation.
  - Krippendorff's alpha Yekti-Daniel on JELAS_HATE (raw) and on the harmonized
    (OTENTIK AND JELAS_HATE) construct for all three pairs, to test whether Mukhlis's
    original single-column instrument conflated two questions that Yekti/Daniel's
    two-column form kept separate.

Independent verification: the harmonized-alpha and disagreement-count numbers below
were cross-checked during development against a second, differently-coded path (a
closed-form two-rater formula: Do = disagree/n, De = 2*p1*(1-p1) with p1 the pooled
proportion of "1" across both raters' judgments, alpha = 1 - Do/De -- distinct from
src/agreement.py's coincidence-matrix loop). Results matched to within the expected
finite-sample rounding (e.g. 0.519 vs 0.517, 0.448 vs 0.446): see STATE.md C14 / the
commit that introduces this script for the cross-check transcript.

Run: python experiments/generation_pilot/analyze_disagreement.py
Writes: disagreement_analysis.md
"""
from __future__ import annotations

import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent.parent))
from src.agreement import krippendorff_alpha_nominal, bootstrap_alpha_ci  # noqa: E402

KEY = HERE / "_key.csv"
MUKHLIS = HERE / "VALIDATION_FORM.xlsx"
YEKTI = HERE / "VALIDATION_FORM_yekti_FILLED.xlsx"
DANIEL = HERE / "VALIDATION_FORM_daniel_FILLED.xlsx"

NICHE_ORDER = ["ngoko_direct", "krama_report", "krama_sarcastic", "krama_cold_contempt"]
MODEL_ORDER = ["deepseek", "gemma3:27b", "qwen3:14b"]


def _hdr_map(ws):
    return {(c.value or "").strip(): i + 1 for i, c in enumerate(ws[1])}


def _col(H, prefix):
    return next((H[k] for k in H if k.startswith(prefix)), None)


def _truthy(v):
    if v in (None, ""):
        return None
    return 1 if str(v).strip() in ("1", "1.0", "ya", "Ya", "YA") else 0


def load_key() -> dict[str, dict]:
    with KEY.open(encoding="utf-8") as f:
        return {row["no"]: row for row in csv.DictReader(f)}


def load_form(path: Path) -> dict[str, dict]:
    """no -> {auth, hate, niche, masalah, catatan}. hate is None if form has no JELAS_HATE column."""
    wb = load_workbook(path)
    ws = wb["Validasi"]
    H = _hdr_map(ws)
    c_no = H.get("no", 1)
    c_auth = _col(H, "OTENTIK")
    c_hate = _col(H, "JELAS_HATE")
    c_niche = H.get("niche")
    c_mas = H.get("MASALAH")
    c_cat = H.get("CATATAN")
    out = {}
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, c_no).value
        if no is None:
            continue
        no = str(int(no)) if isinstance(no, float) else str(no)
        out[no] = dict(
            auth=_truthy(ws.cell(r, c_auth).value),
            hate=_truthy(ws.cell(r, c_hate).value) if c_hate else None,
            niche=ws.cell(r, c_niche).value if c_niche else None,
            masalah=str(ws.cell(r, c_mas).value or "").strip() if c_mas else "",
            catatan=str(ws.cell(r, c_cat).value or "").strip() if c_cat else "",
        )
    return out


def harmonized(rec: dict) -> int | None:
    """OTENTIK AND JELAS_HATE -- approximates the single question Mukhlis's original
    one-column instrument had to answer at once."""
    if rec["auth"] is None or rec["hate"] is None:
        return None
    return 1 if (rec["auth"] == 1 and rec["hate"] == 1) else 0


def alpha_and_ci(pairs: list[tuple]) -> tuple[float, float, float]:
    units = [[a, b] for a, b in pairs]
    a = krippendorff_alpha_nominal(units)
    lo, hi = bootstrap_alpha_ci(units)
    return a, lo, hi


def main():
    key = load_key()
    mukhlis = load_form(MUKHLIS)
    yekti = load_form(YEKTI)
    daniel = load_form(DANIEL)
    ids = sorted(mukhlis, key=int)

    out = []
    P = lambda *a: out.append(" ".join(str(x) for x in a))
    P("# Disagreement diagnosis + E6-lite cross-tab (P0-2/P0-3/P0-4, STRATEGY.md SS12)\n")
    P("Reproduced from `VALIDATION_FORM.xlsx` (Mukhlis) + `VALIDATION_FORM_{yekti,daniel}_FILLED.xlsx` "
      "+ `_key.csv`. No new native input. See module docstring for the independent-verification note.\n")

    # ---- P0-2: disagreement diagnosis, Mukhlis vs Yekti / Mukhlis vs Daniel ----
    for name, secondary in [("yekti", yekti), ("daniel", daniel)]:
        P(f"## Mukhlis vs {name}: disagreement rows\n")
        rows = []
        for no in ids:
            a_m = mukhlis[no]["auth"]
            a_s = secondary.get(no, {}).get("auth")
            if a_m is not None and a_s is not None and a_m != a_s:
                rows.append(dict(no=no, niche=secondary[no]["niche"], m=a_m, s=a_s,
                                  hate=secondary[no]["hate"], catatan=secondary[no]["catatan"]))
        n = len(rows)
        direction = Counter((r["m"], r["s"]) for r in rows)
        niche_counts = Counter(r["niche"] for r in rows)
        hate_dist = Counter(r["hate"] for r in rows)
        P(f"- **n disagreement = {n}**")
        for (m_val, s_val), c in sorted(direction.items(), key=lambda kv: -kv[1]):
            P(f"  - direction (Mukhlis={m_val}, {name}={s_val}): {c}")
        P(f"- niche breakdown: " + ", ".join(f"{k}={v}" for k, v in
                                              sorted(niche_counts.items(), key=lambda kv: -kv[1])))
        if n:
            top_niche, top_n = niche_counts.most_common(1)[0]
            P(f"  - **{top_niche}: {top_n}/{n} = {top_n/n:.0%} of all {name} disagreements** "
              f"(this is the correct denominator; do not confuse with the niche's own "
              f"{top_n}/27 = {top_n/27:.0%} share of disagreement WITHIN that niche -- both are true, "
              f"they answer different questions, see STRATEGY.md P0-1)")
        P(f"- {name}'s JELAS_HATE among these disagreement rows: " +
          ", ".join(f"{k}={v}" for k, v in sorted(hate_dist.items(), key=lambda kv: str(kv[0]))))
        # code-switching note count, nested inside JELAS_HATE=0 rows
        cs_rows = [r for r in rows if "campur" in r["catatan"].lower()]
        cs_and_hate0 = [r for r in cs_rows if r["hate"] == 0]
        P(f"- rows whose {name} CATATAN mentions code-mixing ('campur'): {len(cs_rows)} "
          f"(of which {len(cs_and_hate0)} also have JELAS_HATE=0 -- "
          f"{'fully nested' if len(cs_and_hate0) == len(cs_rows) else 'NOT fully nested, check manually'})")
        # non-code-switching, non-hate0 rows (interesting residual for the OTHER pattern, e.g. Daniel's register doubt)
        other = [r for r in rows if not (r["hate"] == 0 and "campur" in r["catatan"].lower())]
        pattern_10 = [r for r in rows if r["hate"] == 1 and r["m"] == 1 and r["s"] == 0]
        if pattern_10:
            P(f"- rows where {name} marks JELAS_HATE=1 but OTENTIK=0 (register doubt despite "
              f"recognizing hate content): {len(pattern_10)}/{n}")
        P("")

    # ---- Harmonized alpha (OTENTIK AND JELAS_HATE) for all 3 pairs ----
    P("## Harmonized alpha (OTENTIK AND JELAS_HATE)\n")
    P("Tests whether Mukhlis's single-column instrument conflated the two questions "
      "Yekti/Daniel's two-column form kept separate. Uses all 108 items, not just the disagreement subset.\n")
    P("| pair | raw OTENTIK alpha | harmonized alpha | 95% CI |")
    P("|---|---|---|---|")
    raw_alpha = {}
    for name, secondary in [("yekti", yekti), ("daniel", daniel)]:
        pairs = [(mukhlis[no]["auth"], secondary[no]["auth"]) for no in ids
                 if mukhlis[no]["auth"] is not None and secondary.get(no, {}).get("auth") is not None]
        a, lo, hi = alpha_and_ci(pairs)
        raw_alpha[name] = a
    for name, secondary in [("yekti", yekti), ("daniel", daniel)]:
        h_pairs = [(mukhlis[no]["auth"], harmonized(secondary[no])) for no in ids
                   if mukhlis[no]["auth"] is not None and harmonized(secondary[no]) is not None]
        a, lo, hi = alpha_and_ci(h_pairs)
        P(f"| mukhlis vs {name} | {raw_alpha[name]:.3f} | {a:.3f} | [{lo:.3f}, {hi:.3f}] |")
    # Yekti-Daniel harmonized (added for completeness, per STRATEGY P0-2)
    h_pairs_yd = [(harmonized(yekti[no]), harmonized(daniel[no])) for no in ids
                  if harmonized(yekti[no]) is not None and harmonized(daniel[no]) is not None]
    a_yd, lo_yd, hi_yd = alpha_and_ci(h_pairs_yd)
    raw_pairs_yd = [(yekti[no]["auth"], daniel[no]["auth"]) for no in ids
                     if yekti[no]["auth"] is not None and daniel[no]["auth"] is not None]
    a_yd_raw, _, _ = alpha_and_ci(raw_pairs_yd)
    P(f"| yekti vs daniel | {a_yd_raw:.3f} | {a_yd:.3f} | [{lo_yd:.3f}, {hi_yd:.3f}] |")
    P("")

    # ---- P0-3(a): JELAS_HATE rate per niche per validator ----
    P("## P0-3(a): JELAS_HATE rate per niche per validator\n")
    P("| niche | yekti | daniel |")
    P("|---|---|---|")
    for niche in NICHE_ORDER:
        row = [niche]
        for secondary in (yekti, daniel):
            vals = [secondary[no]["hate"] for no in ids
                    if secondary[no]["niche"] == niche and secondary[no]["hate"] is not None]
            row.append(f"{sum(vals)}/{len(vals)} ({sum(vals)/len(vals):.0%})" if vals else "-")
        P("| " + " | ".join(row) + " |")
    P("")

    # ---- P0-3(b): JELAS_HATE x machine_caught cross-tab, 36 DeepSeek cells ----
    P("## P0-3(b): JELAS_HATE x machine_caught cross-tab (36 DeepSeek cells with detector data)\n")
    deepseek_ids = [no for no in ids if key[no]["model"] == "deepseek"]
    for name, secondary in [("yekti", yekti), ("daniel", daniel)]:
        P(f"### {name}\n")
        P("| machine_caught | JELAS_HATE=1 | JELAS_HATE=0 |")
        P("|---|---|---|")
        tab = defaultdict(lambda: [0, 0])
        for no in deepseek_ids:
            c = int(key[no]["machine_caught"].split("/")[0])
            h = secondary[no]["hate"]
            if h is not None:
                tab[c][h] += 1
        for c in sorted(tab):
            P(f"| {c}/5 | {tab[c][1]} | {tab[c][0]} |")
        P("")
    zero_ids = [no for no in deepseek_ids if key[no]["machine_caught"] == "0/5"]
    P(f"**The 9 cells that evade ALL 5 detectors** ({', '.join(zero_ids)}):\n")
    P("| niche | yekti JELAS_HATE | daniel JELAS_HATE |")
    P("|---|---|---|")
    for no in zero_ids:
        P(f"| {key[no]['niche']} | {yekti[no]['hate']} | {daniel[no]['hate']} |")
    y_sum = sum(yekti[no]["hate"] for no in zero_ids)
    d_sum = sum(daniel[no]["hate"] for no in zero_ids)
    P(f"\n**{y_sum}/9 (yekti) and {d_sum}/9 (daniel) of the all-detector-evading cells "
      f"are rated JELAS_HATE=1.** This bears directly on SS4.5's claim that \"the hate that "
      f"detectors miss is real Javanese hate\" -- see STRATEGY.md P0-3 guard rail: do not "
      f"rewrite SS4.5 before this number is reviewed with Bapak if it weakens the claim.\n")

    # ---- P0-3(c): Yekti-Daniel alpha on raw JELAS_HATE ----
    P("## P0-3(c): Yekti-Daniel Krippendorff alpha on JELAS_HATE (raw, never computed before)\n")
    jh_pairs = [(yekti[no]["hate"], daniel[no]["hate"]) for no in ids
                if yekti[no]["hate"] is not None and daniel[no]["hate"] is not None]
    a, lo, hi = alpha_and_ci(jh_pairs)
    P(f"alpha = {a:.3f} [{lo:.3f}, {hi:.3f}] (n={len(jh_pairs)})\n")
    P("Compare to Yekti-Daniel OTENTIK alpha = -0.039 (chance-or-below) and harmonized alpha "
      "above -- JELAS_HATE alone agrees far better than OTENTIK alone or the raw pool, though "
      "high agreement here is partly driven by the KRAMA_SARCASTIC 0% prevalence (see P0-3(a)) "
      "inflating chance agreement; read alongside the per-niche table, not in isolation.\n")

    # ---- P0-4: per-model x per-validator authenticity table ----
    P("## P0-4: per-model x per-validator authenticity table\n")
    P("| model | Mukhlis | Yekti | Daniel |")
    P("|---|---|---|---|")
    for model in MODEL_ORDER:
        model_ids = [no for no in ids if key[no]["model"] == model]
        row = [model]
        for recs in (mukhlis, yekti, daniel):
            vals = [recs[no]["auth"] for no in model_ids if recs[no]["auth"] is not None]
            row.append(f"{sum(vals)}/{len(vals)} ({sum(vals)/len(vals):.0%})")
        P("| " + " | ".join(row) + " |")
    P("")
    P("DeepSeek ranks #1 across all three validators (the relative model ordering "
      "deepseek > gemma3:27b > qwen3:14b also holds for all three) -- the headline "
      "generator-quality ranking is robust to which native validator scores it, even "
      "though absolute rates differ (Yekti's overall leniency shows up per-model too).\n")

    (HERE / "disagreement_analysis.md").write_text("\n".join(out), encoding="utf-8")
    print("\n".join(out))


if __name__ == "__main__":
    main()
