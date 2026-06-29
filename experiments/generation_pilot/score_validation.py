"""Score the native expert's authenticity validation of the generated hate.

Run AFTER Bapak fills VALIDATION_FORM.xlsx (sheet 'Validasi'):
  OTENTIK?  -> 1 (authentic Javanese hate) / 0 (not)
  MASALAH   -> museum_krama | bocor_indonesia | bukan_hate | salah_register | salah_target | formulaic | lainnya
  CATATAN   -> free text

  python experiments/generation_pilot/score_validation.py

Columns are located BY HEADER (robust to the enriched multi-model form). Computes the
questions that matter for the paper:
  - overall authenticity rate of LLM-generated Javanese hate
  - PER-NICHE rate  -> does N3b group-directed krama cold-contempt hold up natively?
  - PER-MODEL rate  -> which generator (deepseek vs gemma3 vs qwen3) is the better source?
  - model x niche   -> the inter-model x register table (FINDINGS §5)
  - per-target rate, failure reasons
  - DETECTOR-EVASION x NATIVE: do detector-evasive items still read as authentic hate?
    (evasive AND authentic = the genuinely hard hate cheap detectors miss = the headline)
  - does the non-native auto-flag heuristic / judge panel predict native rejection?
Writes validation_result.md
"""
from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
FORM = HERE / "VALIDATION_FORM.xlsx"
KEY = HERE / "_key.csv"

NICHE_ORDER = ["ngoko_direct", "krama_report", "krama_sarcastic", "krama_cold_contempt"]


def _hdr_map(ws):
    return {(c.value or "").strip(): i + 1 for i, c in enumerate(ws[1])}


def _truthy(v) -> int:
    return 1 if str(v).strip() in ("1", "1.0", "ya", "Ya", "YA") else 0


def main():
    if not FORM.exists() or not KEY.exists():
        print("Missing VALIDATION_FORM.xlsx or _key.csv. Run rebuild_form.py first.")
        return
    key = {}
    with KEY.open(encoding="utf-8") as f:
        for row in csv.DictReader(f):
            key[str(row["no"])] = row

    wb = load_workbook(FORM)
    ws = wb["Validasi"]
    H = _hdr_map(ws)
    c_no = H.get("no", 1)
    c_auth = next((H[k] for k in H if k.startswith("OTENTIK")), None)
    c_mas = H.get("MASALAH")
    c_cat = H.get("CATATAN")
    if not c_auth:
        print("Could not find OTENTIK column.")
        return

    recs, missing = [], []
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, c_no).value
        if no is None:
            continue
        auth = ws.cell(r, c_auth).value
        masalah = (ws.cell(r, c_mas).value if c_mas else "") or ""
        catatan = (ws.cell(r, c_cat).value if c_cat else "") or ""
        k = key.get(str(int(no)) if isinstance(no, float) else str(no), {})
        if auth in (None, ""):
            missing.append(str(no))
            continue
        recs.append(dict(
            no=str(no), model=k.get("model", "?"), niche=k.get("niche", "?"),
            target=k.get("intended_target", "?"), auth=_truthy(auth),
            masalah=str(masalah).strip(), catatan=str(catatan).strip(),
            machine=k.get("machine_caught", ""), concern=k.get("auto_concern", ""),
            prioritas=k.get("prioritas", "0"), source=k.get("source", ""),
        ))

    out = []
    P = lambda *a: out.append(" ".join(str(x) for x in a))
    P("# Generation pilot — native authenticity validation\n")
    if not recs:
        P("> No rows scored yet. Fill the OTENTIK? column in VALIDATION_FORM.xlsx.")
        (HERE / "validation_result.md").write_text("\n".join(out), encoding="utf-8")
        print("\n".join(out))
        return
    if missing:
        P(f"> NOTE: {len(missing)} rows unrated.\n")

    n = len(recs)
    ok = sum(r["auth"] for r in recs)
    P(f"Scored **{ok}/{n} = {ok/n:.0%}** judged authentic Javanese hate.\n")

    def rate_table(title, keyfn, order=None):
        P(f"## {title}\n")
        P("| group | authentic | rate |")
        P("|---|---|---|")
        buckets = defaultdict(list)
        for r in recs:
            buckets[keyfn(r)].append(r["auth"])
        keys = order if order else sorted(buckets, key=lambda k: -sum(buckets[k]) / len(buckets[k]))
        for k in keys:
            v = buckets.get(k)
            if v:
                P(f"| {k} | {sum(v)}/{len(v)} | {sum(v)/len(v):.0%} |")
        P("")

    rate_table("Per niche (register-pragmatic axis)", lambda r: r["niche"],
               [x for x in NICHE_ORDER if any(r["niche"] == x for r in recs)])
    P("_krama_cold_contempt = the open N3b-group question (can SARA-group cold-contempt krama hate "
      "be generated authentically?)._\n")
    rate_table("Per model (which generator is the better source?)", lambda r: r["model"])

    # model x niche matrix
    models = sorted({r["model"] for r in recs})
    niches = [x for x in NICHE_ORDER if any(r["niche"] == x for r in recs)]
    P("## Model x niche authenticity rate\n")
    P("| model \\ niche | " + " | ".join(niches) + " |")
    P("|" + "---|" * (len(niches) + 1))
    mn = defaultdict(list)
    for r in recs:
        mn[(r["model"], r["niche"])].append(r["auth"])
    for m in models:
        cells = []
        for ni in niches:
            v = mn.get((m, ni))
            cells.append(f"{sum(v)}/{len(v)}" if v else "–")
        P(f"| {m} | " + " | ".join(cells) + " |")
    P("")

    rate_table("Per target (SARA coverage)", lambda r: r["target"])

    # detector-evasion x native authenticity (deepseek only — has machine_caught)
    ev = [r for r in recs if r["machine"] and "/" in r["machine"]]
    if ev:
        P("## Detector-evasion x native authenticity (the headline cross-tab)\n")
        def caught_frac(m):
            c, v = m.split("/")
            return int(c) / int(v) if int(v) else None
        evasive = [r for r in ev if (caught_frac(r["machine"]) or 1) <= 0.5]
        caught = [r for r in ev if (caught_frac(r["machine"]) or 1) > 0.5]
        def rate(rs):
            return f"{sum(x['auth'] for x in rs)}/{len(rs)} ({sum(x['auth'] for x in rs)/len(rs):.0%})" if rs else "–"
        P("| detector verdict | native-authentic rate | meaning |")
        P("|---|---|---|")
        P(f"| evaded by >=half detectors | {rate(evasive)} | authentic+evasive = hate cheap detectors MISS (the point) |")
        P(f"| caught by >half detectors | {rate(caught)} | authentic+caught = detectors already handle these |")
        P("")

    # failure reasons
    fails = [r for r in recs if not r["auth"]]
    P(f"## Failure reasons ({len(fails)} rejected)\n")
    reasons = defaultdict(int)
    for r in fails:
        reasons[r["masalah"] or "(unspecified)"] += 1
    for reason, c in sorted(reasons.items(), key=lambda kv: -kv[1]):
        P(f"- {reason}: {c}")
    if fails:
        P("\nRejected items:")
        for r in fails:
            note = f" — {r['catatan']}" if r["catatan"] else ""
            P(f"- [{r['model']}/{r['niche']}/{r['target']}] {r['masalah']}{note}")

    # auto-flag (judge panel) vs native
    P("\n## QC judge panel vs native verdict\n")
    flagged = [r for r in recs if r["concern"]]
    if flagged:
        tp = sum(1 for r in flagged if not r["auth"])
        P(f"- {len(flagged)} judge-flagged; {tp} natively rejected (precision {tp/len(flagged):.0%}).")
    if fails:
        fr = sum(1 for r in fails if r["concern"])
        P(f"- of {len(fails)} native rejections, {fr} were judge-flagged (recall {fr/len(fails):.0%}).")

    # priority subset
    prio = [r for r in recs if str(r["prioritas"]) == "1"]
    if prio:
        pj = sum(r["auth"] for r in prio)
        P(f"\n## PRIORITAS subset: {pj}/{len(prio)} = {pj/len(prio):.0%} authentic\n")

    P("## Interpretation guide")
    P("- High overall rate => LLM is a viable generator for a register-stratified Javanese hate set.")
    P("- krama_cold_contempt authentic => N3b-group is generable (the uncollectable register synthesized + native-validated).")
    P("- Best model x niche cell => preferred generator per register.")
    P("- Evasive AND authentic => genuinely hard hate the cheap detectors miss = the dataset's reason to exist.")
    P("- Low judge recall => non-native auto-checks miss native-only judgments => confirms the irreducible native-referee role.")

    (HERE / "validation_result.md").write_text("\n".join(out), encoding="utf-8")
    print("\n".join(out))


if __name__ == "__main__":
    main()
