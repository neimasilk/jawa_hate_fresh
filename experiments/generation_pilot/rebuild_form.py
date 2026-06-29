"""Rebuild the native-validation form, enriched for an efficient single pass.

Supersedes review_generated.py's form. Combines every generator (DeepSeek from
generated_pilot.jsonl + local models from generated_multimodel.jsonl) into ONE form
and adds machine-context columns so the native expert can spend his scarce time well:

  - machine_caught : for DeepSeek examples, how many of the 5 production-grade
                     detectors flagged it as hate (from detect_results.jsonl).
                     LOW = detector-evasive = the most interesting items to validate.
  - auto_concern   : QC concerns raised by the non-native judge panel (judge_panel.json),
                     e.g. museum_krama / register_mismatch / formulaic_opener.
  - PRIORITAS      : 1 for the highest-value rows (detector-evasive OR judge-flagged
                     DeepSeek items + a stratified slice of each local model x niche).
                     Bapak does PRIORITAS=1 first; the rest if time permits.

Never clobbers a form that already has authenticity verdicts filled in.
Run: python experiments/generation_pilot/rebuild_form.py
Writes: VALIDATION_FORM.xlsx, _key.csv
"""
from __future__ import annotations

import csv
import json
from collections import defaultdict
from pathlib import Path

HERE = Path(__file__).parent
DEEPSEEK = HERE / "generated_pilot.jsonl"
MULTIMODEL = HERE / "generated_multimodel.jsonl"
DETECT = HERE / "detect_results.jsonl"
JUDGE = HERE / "judge_panel.json"
FORM = HERE / "VALIDATION_FORM.xlsx"
KEY = HERE / "_key.csv"

HEADERS = ["no", "model", "niche", "target", "register", "TEKS", "mekanisme",
           "machine_caught", "auto_concern", "PRIORITAS",
           "OTENTIK? (1=ya,0=tidak)", "MASALAH", "CATATAN"]


def _short_model(m: str) -> str:
    return {"deepseek-v4-pro": "deepseek"}.get(m, m)


def load_examples() -> list[dict]:
    rows = []
    # DeepSeek: gen_no = 1-based line number (matches detect_probe.py)
    if DEEPSEEK.exists():
        for i, line in enumerate(
                (l for l in DEEPSEEK.read_text(encoding="utf-8").splitlines() if l.strip()), 1):
            r = json.loads(line)
            rows.append({
                "model": _short_model(r.get("model", "deepseek-v4-pro")),
                "gen_no": i, "niche": r["niche"], "target": r["intended_target"],
                "register": r.get("register", ""), "text": r.get("text", ""),
                "mekanisme": r.get("mekanisme", ""), "source": "deepseek",
            })
    # local models
    if MULTIMODEL.exists():
        for line in MULTIMODEL.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            r = json.loads(line)
            rows.append({
                "model": _short_model(r.get("model", "?")),
                "gen_no": None, "niche": r["niche"], "target": r["intended_target"],
                "register": r.get("register", ""), "text": r.get("text", ""),
                "mekanisme": r.get("mekanisme", ""), "source": "local",
            })
    return rows


def load_detect() -> dict[int, str]:
    """gen_no -> 'caught/valid' across detectors (latest verdict per detector)."""
    if not DETECT.exists():
        return {}
    latest = {}
    for line in DETECT.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            r = json.loads(line)
        except json.JSONDecodeError:
            continue
        latest[(r["gen_no"], r["detector"])] = r.get("hate")
    agg = defaultdict(lambda: [0, 0])  # caught, valid
    for (gen_no, _det), hate in latest.items():
        if hate is None:
            continue
        agg[gen_no][1] += 1
        if hate:
            agg[gen_no][0] += 1
    return {no: f"{c}/{v}" for no, (c, v) in agg.items()}


def load_judge() -> dict[int, dict]:
    """no -> {concerns:int, issues:[...]} from the QC judge panel."""
    if not JUDGE.exists():
        return {}
    data = json.loads(JUDGE.read_text(encoding="utf-8"))
    return {int(k): v for k, v in data.get("by_no", {}).items()}


def existing_filled() -> bool:
    if not FORM.exists():
        return False
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False
    wb = load_workbook(FORM)
    if "Validasi" not in wb.sheetnames:
        return False
    ws = wb["Validasi"]
    hdr = [c.value for c in ws[1]]
    if "OTENTIK? (1=ya,0=tidak)" not in hdr:
        return False
    gcol = hdr.index("OTENTIK? (1=ya,0=tidak)") + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, gcol).value not in (None, ""):
            return True
    return False


def main():
    if existing_filled():
        print("ABORT: VALIDATION_FORM.xlsx already has authenticity verdicts filled. "
              "Refusing to overwrite. Score it first or move it aside.")
        return

    rows = load_examples()
    if not rows:
        print("No generated examples found.")
        return
    detect = load_detect()
    judge = load_judge()

    # assign global running numbers + machine/judge columns + PRIORITAS
    local_seen = set()  # (model, niche) -> mark first as priority slice
    enriched = []
    for idx, r in enumerate(rows, 1):
        machine = detect.get(r["gen_no"], "") if r["gen_no"] else ""
        j = judge.get(r["gen_no"]) if r["gen_no"] else None
        concerns = j["concerns"] if j else 0
        issues = ";".join(sorted(set(j["issues"]))) if j else ""

        prioritas = 0
        if r["source"] == "deepseek":
            caught = None
            if machine:
                c, v = machine.split("/")
                caught = (int(c), int(v))
            evasive = caught is not None and caught[1] > 0 and caught[0] <= caught[1] / 2
            flagged = concerns >= 1
            if evasive or flagged:
                prioritas = 1
        else:  # local: one row per (model, niche) as a comparison slice
            mk = (r["model"], r["niche"])
            if mk not in local_seen:
                local_seen.add(mk)
                prioritas = 1

        enriched.append({**r, "no": idx, "machine_caught": machine,
                         "auto_concern": issues, "prioritas": prioritas})

    # --- _key.csv ---
    with KEY.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["no", "model", "niche", "intended_target", "register", "text",
                    "mekanisme", "machine_caught", "auto_concern", "prioritas", "source"])
        for r in enriched:
            w.writerow([r["no"], r["model"], r["niche"], r["target"], r["register"],
                        r["text"], r["mekanisme"], r["machine_caught"], r["auto_concern"],
                        r["prioritas"], r["source"]])

    # --- xlsx ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl missing; wrote _key.csv only.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Validasi"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
    prio_fill = PatternFill("solid", fgColor="FFF2CC")
    for r in enriched:
        ws.append([r["no"], r["model"], r["niche"], r["target"], r["register"], r["text"],
                   r["mekanisme"], r["machine_caught"], r["auto_concern"],
                   r["prioritas"] or "", "", "", ""])
        if r["prioritas"]:
            for c in range(1, len(HEADERS) + 1):
                ws.cell(ws.max_row, c).fill = prio_fill
    widths = {"A": 5, "B": 10, "C": 20, "D": 16, "E": 9, "F": 58, "G": 44,
              "H": 9, "I": 22, "J": 9, "K": 16, "L": 14, "M": 26}
    for col, wd in widths.items():
        ws.column_dimensions[col].width = wd
    for rr in range(2, ws.max_row + 1):
        ws.cell(rr, 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(rr, 7).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    ins = wb.create_sheet("Petunjuk")
    n_prio = sum(r["prioritas"] for r in enriched)
    guide = [
        ["VALIDASI KEASLIAN — ujaran kebencian Jawa hasil-generate LLM (multi-model)"],
        [""],
        [f"Total {len(enriched)} contoh. Baris kuning (PRIORITAS=1, {n_prio} baris) = isi DULU."],
        ["PRIORITAS = contoh yang LOLOS deteksi mesin (paling menarik utk dataset) atau di-flag panel-QC,"],
        ["plus 1 contoh per (model lokal x niche) untuk perbandingan antar-model. Sisanya kalau sempat."],
        [""],
        ["Untuk tiap baris isi 3 kolom kanan:"],
        ["K  OTENTIK?  -> 1 jika terdengar seperti hate Jawa ASLI (orang Jawa beneran bisa nulis ini); 0 jika tidak."],
        ["L  MASALAH   -> kalau 0 (atau ada catatan), pilih: museum_krama | bocor_indonesia | bukan_hate |"],
        ["                salah_register | salah_target | formulaic | lainnya"],
        ["M  CATATAN   -> opsional: koreksi, versi lebih otentik, dialek, dll."],
        [""],
        ["Kolom konteks (read-only, untuk bantu menilai — JANGAN diubah):"],
        ["H  machine_caught = berapa dari 5 detektor mesin menandai ini 'hate' (mis. 2/5). Kecil = lolos deteksi."],
        ["I  auto_concern   = flag dari panel-QC LLM non-native (museum_krama/register_mismatch/formulaic/...)."],
        [""],
        ["Fokus: KEASLIAN, bukan apakah Anda setuju isinya. Ini data riset DETEKSI (defensif)."],
    ]
    for row in guide:
        ins.append(row)
    ins.column_dimensions["A"].width = 110

    wb.save(FORM)

    by_model = defaultdict(int)
    for r in enriched:
        by_model[r["model"]] += 1
    print(f"Rebuilt form: {len(enriched)} examples "
          f"({', '.join(f'{m}={n}' for m, n in by_model.items())}); {n_prio} PRIORITAS.")
    print(f"  machine_caught populated: {sum(1 for r in enriched if r['machine_caught'])}")
    print(f"  judge concerns populated: {sum(1 for r in enriched if r['auto_concern'])}")
    print(f"  -> {FORM.name}, {KEY.name}")


if __name__ == "__main__":
    main()
