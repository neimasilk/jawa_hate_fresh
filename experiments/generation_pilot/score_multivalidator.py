"""Score inter-rater reliability across native validators (E1, STRATEGY.md).

Run AFTER one or more of VALIDATION_FORM_yekti.xlsx / VALIDATION_FORM_daniel.xlsx have
been filled in (see build_multivalidator_forms.py). Combines each filled form with
Mukhlis's already-scored VALIDATION_FORM.xlsx to compute:

  - per-validator authenticity rate (comparable to Mukhlis's single-evaluator 55%)
  - pairwise raw agreement + Krippendorff's alpha on OTENTIK (canonical formula,
    src/agreement.py, D17) for every pair of validators who have answered
  - the OTENTIK vs JELAS_HATE cross-tab per new validator (does authenticity track
    hate-clarity, or do they diverge -- e.g. authentic-but-not-hate cases like the
    paper's example #21)

This does not change what's already in the paper (Mukhlis's 97%/56%/11%, 55% overall
figures stay as the single-evaluator baseline); it adds inter-rater numbers that were
previously missing (Limitation #1 in draft_jinita.md).

Run: python experiments/generation_pilot/score_multivalidator.py
Writes: multivalidator_result.md
"""
from __future__ import annotations

import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE.parent.parent))
from src.agreement import krippendorff_alpha_nominal, bootstrap_alpha_ci  # noqa: E402

PRIMARY = ("mukhlis", HERE / "VALIDATION_FORM.xlsx")
SECONDARY = [
    ("yekti", HERE / "VALIDATION_FORM_yekti_FILLED.xlsx"),
    ("daniel", HERE / "VALIDATION_FORM_daniel_FILLED.xlsx"),
]


def _hdr_map(ws):
    return {(c.value or "").strip(): i + 1 for i, c in enumerate(ws[1])}


def _truthy(v):
    if v in (None, ""):
        return None
    return 1 if str(v).strip() in ("1", "1.0", "ya", "Ya", "YA") else 0


def load_form(path: Path) -> dict[str, dict] | None:
    """no -> {auth, hate_clear, masalah, catatan}. None if file missing/empty."""
    if not path.exists():
        return None
    wb = load_workbook(path)
    if "Validasi" not in wb.sheetnames:
        return None
    ws = wb["Validasi"]
    H = _hdr_map(ws)
    c_no = H.get("no", 1)
    c_auth = next((H[k] for k in H if k.startswith("OTENTIK")), None)
    c_hate = next((H[k] for k in H if k.startswith("JELAS_HATE")), None)
    c_mas = H.get("MASALAH")
    c_cat = H.get("CATATAN")
    if not c_auth:
        return None
    out = {}
    any_filled = False
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, c_no).value
        if no is None:
            continue
        no = str(int(no)) if isinstance(no, float) else str(no)
        auth = _truthy(ws.cell(r, c_auth).value)
        hate_clear = _truthy(ws.cell(r, c_hate).value) if c_hate else None
        if auth is not None:
            any_filled = True
        out[no] = dict(
            auth=auth, hate_clear=hate_clear,
            masalah=str(ws.cell(r, c_mas).value or "").strip() if c_mas else "",
            catatan=str(ws.cell(r, c_cat).value or "").strip() if c_cat else "",
        )
    return out if any_filled else None


def main():
    primary_name, primary_path = PRIMARY
    primary = load_form(primary_path)
    if not primary:
        print(f"Missing or empty {primary_path.name}. This should already be scored "
              f"(see validation_result.md) -- check the file exists.")
        return

    validators = {primary_name: primary}
    for name, path in SECONDARY:
        f = load_form(path)
        if f:
            validators[name] = f
        else:
            print(f"({name}: {path.name} not found or not yet filled -- skipping)")

    out = []
    P = lambda *a: out.append(" ".join(str(x) for x in a))
    P("# Inter-rater reliability — native authenticity validation\n")
    P(f"Validators with data: {', '.join(validators)}\n")

    if len(validators) < 2:
        P("> Only one validator scored so far. Run build_multivalidator_forms.py, "
          "have Yekti/Daniel fill VALIDATION_FORM_<name>.xlsx blind, then re-run this script.")
        (HERE / "multivalidator_result.md").write_text("\n".join(out), encoding="utf-8")
        print("\n".join(out))
        return

    # per-validator authenticity rate
    P("## Per-validator authenticity rate\n")
    P("| validator | authentic | rate |")
    P("|---|---|---|")
    for name, recs in validators.items():
        rated = [r["auth"] for r in recs.values() if r["auth"] is not None]
        if rated:
            P(f"| {name} | {sum(rated)}/{len(rated)} | {sum(rated)/len(rated):.0%} |")
    P("")

    # pairwise agreement + Krippendorff alpha on OTENTIK
    P("## Pairwise inter-rater reliability (OTENTIK)\n")
    P("| pair | n shared | raw agreement | Krippendorff alpha | 95% CI |")
    P("|---|---|---|---|---|")
    names = list(validators)
    all_ids = set()
    for recs in validators.values():
        all_ids |= set(recs)
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            a, b = names[i], names[j]
            shared = [no for no in all_ids
                      if validators[a].get(no, {}).get("auth") is not None
                      and validators[b].get(no, {}).get("auth") is not None]
            if not shared:
                continue
            units = [[validators[a][no]["auth"], validators[b][no]["auth"]] for no in shared]
            agree = sum(1 for u in units if u[0] == u[1])
            alpha = krippendorff_alpha_nominal(units)
            lo, hi = bootstrap_alpha_ci(units)
            P(f"| {a} vs {b} | {len(shared)} | {agree}/{len(shared)} ({agree/len(shared):.0%}) "
              f"| {alpha:.3f} | [{lo:.3f}, {hi:.3f}] |")
    P("")

    # 3-way alpha if all three present
    if len(names) >= 3:
        shared3 = [no for no in all_ids
                   if all(validators[n].get(no, {}).get("auth") is not None for n in names)]
        if shared3:
            units = [[validators[n][no]["auth"] for n in names] for no in shared3]
            alpha = krippendorff_alpha_nominal(units)
            lo, hi = bootstrap_alpha_ci(units)
            P(f"**All-{len(names)}-validator alpha** (n={len(shared3)}): "
              f"{alpha:.3f} [{lo:.3f}, {hi:.3f}]\n")

    # OTENTIK vs JELAS_HATE cross-tab, per secondary validator (primary form has no such column)
    for name, recs in validators.items():
        if name == primary_name:
            continue
        both = [(r["auth"], r["hate_clear"]) for r in recs.values()
                if r["auth"] is not None and r["hate_clear"] is not None]
        if not both:
            continue
        P(f"## {name}: OTENTIK vs JELAS_HATE cross-tab\n")
        P("| | JELAS_HATE=1 | JELAS_HATE=0 |")
        P("|---|---|---|")
        for a_val, a_label in [(1, "OTENTIK=1"), (0, "OTENTIK=0")]:
            row = [str(sum(1 for auth, hate in both if auth == a_val and hate == h)) for h in (1, 0)]
            P(f"| {a_label} | {row[0]} | {row[1]} |")
        divergent = sum(1 for auth, hate in both if auth != hate)
        P(f"\n{divergent}/{len(both)} items diverge (authentic-but-not-clearly-hate, or vice versa) "
          f"— these are the cases where the single-dimension OTENTIK column in the original form "
          f"would have been ambiguous.\n")

    P("## Interpretation guide")
    P("- High pairwise agreement / alpha => the 55% single-evaluator authenticity rate "
      "generalizes across native speakers, closing Limitation #1 in the paper.")
    P("- Low agreement => authenticity judgment is itself noisy/subjective for some items; "
      "report per-validator rates separately rather than a pooled headline number.")
    P("- Large OTENTIK/JELAS_HATE divergence => the two dimensions genuinely need to stay "
      "separate in future validation rounds (don't collapse back into one column).")

    (HERE / "multivalidator_result.md").write_text("\n".join(out), encoding="utf-8")
    print("\n".join(out))


if __name__ == "__main__":
    main()
