"""Review generated Javanese hate + build the native-validation form.

Run AFTER generate.py:
  python experiments/generation_pilot/review_generated.py

Does two things:
  1. NON-NATIVE auto-sanity flags (heuristic, NOT ground truth — for triage only):
       - museum_krama : contains archaic/literary krama words a living speaker avoids
       - indo_leak    : no distinctively-Javanese token found (maybe plain Indonesian)
       - register_mismatch : krama niche but self-reported register != krama
     plus the cell-coverage matrix (niche x target).
  2. Emits the AUTHENTICITY validation form for the native expert (Bapak), in the
     same style as experiments/expert_spotcheck: VALIDATION_FORM.xlsx + _key.csv.
     The native rates each generated example: is it AUTHENTIC Javanese hate?

Writes: review_report.md, VALIDATION_FORM.xlsx, _key.csv
"""
from __future__ import annotations

import csv
import json
import re
from collections import defaultdict
from pathlib import Path

HERE = Path(__file__).parent
GEN = HERE / "generated_pilot.jsonl"

# --- heuristics (clearly NOT authoritative; native verdict overrides) ----------

# Literary / fossil krama that an educated native does not actively SPEAK
# (seeded from FINDINGS.md "museum krama": prayogi, dedunung, kitha).
MUSEUM_KRAMA = [
    "prayogi", "dedunung", "kitha", "pawiyatan", "paramasastra",
    "nupiksani", "winastan", "memetri", "pepundhen", "pepoyan",
]

# Distinctively-Javanese tokens (ngoko + krama). Absence => possible Indonesia leak.
JAWA_MARKERS = [
    # ngoko
    "ora", "ojo", "aja", "iki", "iku", "kuwi", "sing", "karo", "wong", "kowe",
    "aku", "wae", "rasah", "mlebu", "metu", "dhewe", "ngono", "ngene", "opo", "apa",
    "kabeh", "gawe", "isin", "jancuk", "asu", "cuk", "raiso", "iso", "nang", "neng",
    # krama
    "panjenengan", "kula", "mboten", "menika", "sampun", "dhateng", "tiyang",
    "inggih", "sanget", "wonten", "kersa", "ngaten", "mugi", "tiyangipun",
]

WORD = re.compile(r"[a-zA-Z']+")


def flags(text: str, niche: str, register: str) -> list[str]:
    low = text.lower()
    out = []
    if any(w in low for w in MUSEUM_KRAMA):
        out.append("museum_krama")
    toks = set(WORD.findall(low))
    if not (toks & set(JAWA_MARKERS)):
        out.append("indo_leak?")
    if niche.startswith("krama") and register and "krama" not in register.lower():
        out.append("register_mismatch")
    return out


def main():
    if not GEN.exists():
        print(f"Missing {GEN}. Run generate.py first.")
        return
    rows = [json.loads(l) for l in GEN.read_text(encoding="utf-8").splitlines() if l.strip()]

    niches, targets = [], []
    for r in rows:
        if r["niche"] not in niches:
            niches.append(r["niche"])
        if r["intended_target"] not in targets:
            targets.append(r["intended_target"])

    # coverage matrix
    cell = defaultdict(int)
    for r in rows:
        cell[(r["niche"], r["intended_target"])] += 1

    # attach auto-flags
    for r in rows:
        r["_flags"] = flags(r.get("text", ""), r["niche"], r.get("register", ""))

    # --- review_report.md ---
    out = []
    P = lambda *a: out.append(" ".join(str(x) for x in a))
    P("# Generation pilot — auto-review (non-native triage)\n")
    P(f"Total generated: **{len(rows)}** examples across "
      f"{len(niches)} niches x {len(targets)} targets.\n")

    P("## Cell coverage (niche x target)\n")
    P("| niche \\ target | " + " | ".join(t.replace('_', ' ') for t in targets) + " |")
    P("|" + "---|" * (len(targets) + 1))
    for n in niches:
        cells = " | ".join(str(cell[(n, t)]) if cell[(n, t)] else "·" for t in targets)
        P(f"| {n} | {cells} |")
    P("")

    flagged = [r for r in rows if r["_flags"]]
    P(f"## Auto-flags (heuristic — native confirms): {len(flagged)}/{len(rows)} flagged\n")
    for tag in ("museum_krama", "indo_leak?", "register_mismatch"):
        n = sum(1 for r in rows if tag in r["_flags"])
        P(f"- `{tag}`: {n}")
    P("")

    P("## All examples by niche\n")
    for n in niches:
        P(f"### {n}\n")
        for r in [x for x in rows if x["niche"] == n]:
            fl = (" ⚠️ " + ",".join(r["_flags"])) if r["_flags"] else ""
            P(f"- **[{r['intended_target']}]** ({r['register']}/{r['severity']}/{r['form']}){fl}")
            P(f"  - {r['text']}")
            P(f"  - _mek:_ {r['mekanisme']}")
        P("")
    (HERE / "review_report.md").write_text("\n".join(out), encoding="utf-8")

    # --- _key.csv (metadata + auto-flags, for later scoring) ---
    with (HERE / "_key.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["no", "niche", "intended_target", "register", "severity",
                    "form", "text", "mekanisme", "auto_flags"])
        for i, r in enumerate(rows, 1):
            w.writerow([i, r["niche"], r["intended_target"], r.get("register", ""),
                        r.get("severity", ""), r.get("form", ""), r.get("text", ""),
                        r.get("mekanisme", ""), ";".join(r["_flags"])])

    # --- VALIDATION_FORM.xlsx for the native expert ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl not installed; skipped xlsx (CSV form below).")
        _csv_form(rows)
        print("\n".join(out[:40]))
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Validasi"
    headers = ["no", "niche", "target", "register", "TEKS", "mekanisme",
               "OTENTIK? (1=ya,0=tidak)", "MASALAH", "CATATAN"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
    for i, r in enumerate(rows, 1):
        ws.append([i, r["niche"], r["intended_target"], r.get("register", ""),
                   r.get("text", ""), r.get("mekanisme", ""), "", "", ""])
    widths = {"A": 5, "B": 22, "C": 22, "D": 9, "E": 60, "F": 50,
              "G": 12, "H": 18, "I": 30}
    for col, wd in widths.items():
        ws.column_dimensions[col].width = wd
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 5).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # instruction sheet
    ins = wb.create_sheet("Petunjuk")
    guide = [
        ["VALIDASI KEASLIAN — ujaran kebencian Jawa hasil-generate LLM"],
        [""],
        ["Untuk tiap baris di sheet 'Validasi', isi 3 kolom:"],
        ["G  OTENTIK?  -> 1 jika teks terdengar seperti ujaran kebencian Jawa ASLI"],
        ["              (orang Jawa sungguhan bisa menulis/mengucapkan ini); 0 jika tidak."],
        ["H  MASALAH   -> jika 0 (atau ada catatan), pilih salah satu:"],
        ["     museum_krama   = krama benar tapi kuno/sastrawi, tak dipakai sehari-hari"],
        ["     bocor_indonesia = sebenarnya Bahasa Indonesia, bukan Jawa"],
        ["     bukan_hate     = tidak benar-benar menyerang kelompok / tidak menyakitkan"],
        ["     salah_register = register tidak sesuai (mis. diminta krama tapi ngoko)"],
        ["     lainnya        = jelaskan di CATATAN"],
        ["I  CATATAN   -> opsional: koreksi, versi yang lebih otentik, dialek, dll."],
        [""],
        ["Fokus: KEASLIAN, bukan apakah Anda setuju isinya. Ini data riset deteksi."],
    ]
    for row in guide:
        ins.append(row)
    ins.column_dimensions["A"].width = 90

    wb.save(HERE / "VALIDATION_FORM.xlsx")

    print(f"Reviewed {len(rows)} examples.")
    print(f"  auto-flagged: {len(flagged)} "
          f"({', '.join(t+'='+str(sum(1 for r in rows if t in r['_flags'])) for t in ('museum_krama','indo_leak?','register_mismatch'))})")
    print(f"  -> review_report.md, _key.csv, VALIDATION_FORM.xlsx")
    missing = [(n, t) for n in niches for t in targets if not cell[(n, t)]]
    if missing:
        print(f"  missing cells ({len(missing)}): {missing}")


def _csv_form(rows):
    with (HERE / "VALIDATION_FORM.csv").open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["no", "niche", "target", "register", "TEKS", "mekanisme",
                    "OTENTIK_1_0", "MASALAH", "CATATAN"])
        for i, r in enumerate(rows, 1):
            w.writerow([i, r["niche"], r["intended_target"], r.get("register", ""),
                        r.get("text", ""), r.get("mekanisme", ""), "", "", ""])


if __name__ == "__main__":
    main()
