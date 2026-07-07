"""Pre-fill the multi-validator forms with an LLM (Claude) DRAFT — NOT native validation.

CONTEXT (read before judging this file):
  - VALIDATION_FORM_yekti.xlsx / VALIDATION_FORM_daniel.xlsx were built BLANK by
    build_multivalidator_forms.py so that two NATIVE Javanese speakers (Yekti, Daniel)
    judge the same 108 items independently of the first author (Mukhlis) -> inter-rater
    reliability. That is the whole point of these forms.
  - On 2026-07-06 the project owner (Mukhlis) EXPLICITLY instructed Claude to pre-fill the
    judgment columns as an LLM draft so the human validators can validate/correct rather
    than start from blank (trading some anchoring bias for speed).
  - This is therefore an LLM-PRE-ANNOTATION + human-validation workflow, NOT a blind
    native rating. Any inter-rater alpha computed on these forms afterwards must be
    reported in the paper as "LLM-pre-annotated, human-validated", NOT "independent blind
    native validation" (that would be misleading).

HOW THE DRAFT WAS PRODUCED (independence guarantee):
  - Judgments are Claude's OWN reading of each text + the helper columns
    (machine_caught, auto_concern) + documented MODEL-LEVEL patterns only
    (e.g. "qwen3 krama tends to leak Indonesian", "kacandran is not a real Javanese
    word"). These are aggregate patterns, fair game.
  - Claude did NOT consult Mukhlis's per-item answers (validation_result.md per-item
    breakdown / the filled VALIDATION_FORM.xlsx OTENTIK column) when judging. Copying
    those would manufacture fake agreement and is a hard line that was NOT crossed.

TWO-COLUMN DESIGN NOTE:
  The new forms split the old single OTENTIK into OTENTIK? (authentic Javanese in the
  intended register) and JELAS_HATE? (clearly attacks a group identity). This separates
  "authentic but not clearly hate" (e.g. #21, sincere-sounding praise of santri piety)
  from "not authentic". The draft reflects that split: DeepSeek #21/#26 are OTENTIK=1
  but JELAS_HATE=0.

Run: python experiments/generation_pilot/_fill_llm_draft.py
Fills both VALIDATION_FORM_<name>.xlsx in place (blanks already backed up to
*_BLANK_BACKUP.xlsx). Marks LLM-origin in the CATATAN column and adds a warning banner to
each file's Petunjuk sheet.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

HERE = Path(__file__).parent
FILES = [HERE / "VALIDATION_FORM_daniel.xlsx", HERE / "VALIDATION_FORM_yekti.xlsx"]

# no -> (OTENTIK, JELAS_HATE, MASALAH, CATATAN).  MASALAH only when OTENTIK==0.
J = {
    # ---- DeepSeek (1-36) ---- ngoko_direct 1-9: authentic ngoko, clearly hate
    1: (1, 1, "", "LLM draft: ngoko vulgar otentik + code-mix realistis"),
    2: (1, 1, "", "LLM draft: ngoko otentik ('jancuk' Surabaya)"),
    3: (1, 1, "", "LLM draft: ngoko otentik"),
    4: (1, 1, "", "LLM draft: ngoko otentik"),
    5: (1, 1, "", "LLM draft: ngoko otentik"),
    6: (1, 1, "", "LLM draft: ngoko otentik ('lonte')"),
    7: (1, 1, "", "LLM draft: ngoko otentik"),
    8: (1, 1, "", "LLM draft: ngoko otentik"),
    9: (1, 1, "", "LLM draft: ngoko otentik + code-mix Indo ringan"),
    # krama_report 10-18: authentic krama (formulaic opener), clearly ngrasani a group
    10: (1, 1, "", "LLM draft: krama otentik (opener formulaic) [PRIORITAS]"),
    11: (1, 1, "", "LLM draft: campur krama+Indo, realistis"),
    12: (1, 1, "", "LLM draft: krama otentik"),
    13: (1, 1, "", "LLM draft: krama otentik [PRIORITAS]"),
    14: (1, 1, "", "LLM draft: krama otentik [PRIORITAS]"),
    15: (1, 1, "", "LLM draft: krama otentik"),
    16: (1, 1, "", "LLM draft: krama otentik"),
    17: (1, 1, "", "LLM draft: krama otentik; batas-tipis politik [PRIORITAS, lolos 0/5]"),
    18: (1, 1, "", "LLM draft: krama otentik [PRIORITAS]"),
    # krama_sarcastic 19-26: authentic krama+irony; #21 & #26 read as sincere praise
    19: (1, 1, "", "LLM draft: krama+ironi otentik; hate halus [PRIORITAS, 1/5]"),
    20: (1, 1, "", "LLM draft: krama+ironi otentik [PRIORITAS, 0/5]"),
    21: (1, 0, "", "LLM draft: krama otentik, TAPI ironi sangat halus - terbaca pujian tulus santri [PRIORITAS, 0/5]. OTENTIK & JELAS_HATE sengaja beda arah."),
    22: (1, 1, "", "LLM draft: krama+ironi otentik [PRIORITAS, 0/5]"),
    23: (1, 1, "", "LLM draft: krama+ironi otentik [PRIORITAS, 1/5]"),
    24: (1, 1, "", "LLM draft: krama+ironi otentik [PRIORITAS, 0/5]"),
    25: (1, 1, "", "LLM draft: campur, ironi otentik [PRIORITAS, 0/5]"),
    26: (1, 0, "", "LLM draft: krama otentik, TAPI ironi ambigu - target serangan tak jelas, terbaca pujian Mataraman [PRIORITAS, 0/5; auto:not_hateful]"),
    # krama_cold_contempt 27-35: authentic krama, recurring 'isin' device
    27: (1, 1, "", "LLM draft: krama cold-contempt otentik (device 'isin')"),
    28: (1, 1, "", "LLM draft: krama otentik [PRIORITAS]"),
    29: (1, 1, "", "LLM draft: krama otentik [PRIORITAS]"),
    30: (1, 1, "", "LLM draft: krama otentik [PRIORITAS, 1/5]"),
    31: (1, 1, "", "LLM draft: krama otentik [PRIORITAS, 4/5]"),
    32: (1, 1, "", "LLM draft: krama otentik"),
    33: (1, 1, "", "LLM draft: krama otentik"),
    34: (1, 1, "", "LLM draft: krama otentik [PRIORITAS, 0/5]"),
    35: (1, 1, "", "LLM draft: krama otentik"),
    # 36: extra deepseek krama_sarcastic x suku_arab (sophisticated pasemon)
    36: (1, 1, "", "LLM draft: krama+Arab pasemon otentik & sophisticated [PRIORITAS, 0/5]"),

    # ---- Gemma3:27b (37-72) ---- ngoko_direct 37-45: authentic ngoko
    37: (1, 1, "", "LLM draft: ngoko otentik [PRIORITAS]"),
    38: (1, 1, "", "LLM draft: ngoko otentik"),
    39: (1, 1, "", "LLM draft: ngoko otentik"),
    40: (1, 1, "", "LLM draft: ngoko otentik"),
    41: (1, 1, "", "LLM draft: ngoko otentik"),
    42: (1, 1, "", "LLM draft: ngoko otentik"),
    43: (1, 1, "", "LLM draft: ngoko otentik"),
    44: (1, 1, "", "LLM draft: ngoko otentik"),
    45: (1, 1, "", "LLM draft: ngoko otentik"),
    # krama_report 46-54: krama opener + Indonesian code-mix (judged authentic-ish; native may rate stricter)
    46: (1, 1, "", "LLM draft: krama report, code-mix Indo ringan ('oleh pencerahan/kadang-kadang') [PRIORITAS]. Native mungkin rate OTENTIK=0 - cek khusus."),
    47: (1, 1, "", "LLM draft: campur, realistis"),
    48: (1, 1, "", "LLM draft: krama report, code-mix Indo ringan"),
    49: (1, 1, "", "LLM draft: krama report, code-mix Indo ringan"),
    50: (1, 1, "", "LLM draft: campur, realistis"),
    51: (1, 1, "", "LLM draft: campur, realistis"),
    52: (1, 1, "", "LLM draft: krama report, code-mix Indo ringan"),
    53: (1, 1, "", "LLM draft: campur, realistis"),
    54: (1, 1, "", "LLM draft: campur, realistis"),
    # krama_sarcastic 55-63: heavy Indonesian leakage (all fail authenticity)
    55: (0, 1, "bocor_indonesia", "LLM draft: bocor Indonesia berat ('panutan sejati/sukses selalu/meskipun') [PRIORITAS]"),
    56: (0, 1, "bocor_indonesia", "LLM draft: mayoritas Indonesia ('rezekine lancar/Mungkin karena...ya?')"),
    57: (0, 1, "bocor_indonesia", "LLM draft: mayoritas Indonesia ('ilmu agamanya/kami orang Jawa ini')"),
    58: (0, 1, "bocor_indonesia", "LLM draft: nyaris Indonesia murni"),
    59: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni ('semangatnya patut diacungi jempol')"),
    60: (0, 1, "bocor_indonesia", "LLM draft: mayoritas Indonesia ('tampil anggun dan mempesona')"),
    61: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni ('mengekspresikan diri')"),
    62: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni ('melayani rakyat')"),
    63: (0, 1, "bocor_indonesia", "LLM draft: mayoritas Indonesia ('berani dan tegas')"),
    # krama_cold_contempt 64-72: some pure ngoko (salah_register), some real krama
    64: (0, 1, "salah_register", "LLM draft: ngoko murni ('Lha iki/ora ngerti/ndak iso'), bukan krama [PRIORITAS]"),
    65: (1, 1, "", "LLM draft: krama opener otentik + ekor Indo (campur)"),
    66: (1, 1, "", "LLM draft: krama otentik ('Kula namung nyebutaken/dipunlunturi')"),
    67: (1, 1, "", "LLM draft: krama otentik ('kula nyuwun pangapunten/aturanipun')"),
    68: (0, 1, "salah_register", "LLM draft: ngoko opener ('Lha iki liyane maneh') + Indo, bukan krama"),
    69: (1, 1, "", "LLM draft: krama otentik ('kula nyuwun pangapunten/polahipun')"),
    70: (0, 1, "salah_register", "LLM draft: ngoko opener ('Lha iki liyane maneh'), bukan krama"),
    71: (1, 1, "", "LLM draft: krama opener otentik + ekor Indo (campur)"),
    72: (0, 1, "salah_register", "LLM draft: ngoko murni ('wong Arek kuwi polah-polah e/ndak iso diatur'), bukan krama"),

    # ---- Qwen3:14b (73-108) ---- ngoko_direct 73-81: authentic ngoko minus 'kacandran' hallucination
    73: (0, 1, "lainnya", "LLM draft: ngoko kasar otentik TAPI 'kacandran' bukan kata Jawa (halusinasi) [PRIORITAS]"),
    74: (1, 1, "", "LLM draft: ngoko otentik (code-mix 'kekacauan' realistis)"),
    75: (1, 1, "", "LLM draft: ngoko otentik ('jancuk')"),
    76: (1, 1, "", "LLM draft: ngoko otentik"),
    77: (1, 1, "", "LLM draft: ngoko otentik ('asu/jancuk')"),
    78: (0, 1, "lainnya", "LLM draft: 'kacandran' bukan kata Jawa + template formulaic"),
    79: (0, 1, "lainnya", "LLM draft: 'kacandran' + 'anjing' (Indo) bukan Jawa"),
    80: (0, 1, "lainnya", "LLM draft: 'kacandran' + 'anjing' (Indo)"),
    81: (0, 1, "lainnya", "LLM draft: 'kacandran' bukan kata Jawa"),
    # krama_report 82-90: krama opener glued to pure Indonesian + 'kacandran'
    82: (0, 1, "bocor_indonesia", "LLM draft: krama opener TAPI 'walaupun aku ngerti mereka nggawe bisnis' = Indonesia murni + 'kacandran' [PRIORITAS]"),
    83: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    84: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    85: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    86: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    87: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    88: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    89: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    90: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni + 'kacandran'"),
    # krama_sarcastic 91-99: ngoko+Indonesian, not krama
    91: (0, 1, "salah_register", "LLM draft: ngoko+Indo ('Wong X iki kok bisa...sing nggak ngerti'), bukan krama [PRIORITAS]"),
    92: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    93: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    94: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    95: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    96: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    97: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    98: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    99: (0, 1, "salah_register", "LLM draft: ngoko+Indo, bukan krama"),
    # krama_cold_contempt 100-108: pure Indonesian
    100: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni ('nggak pernah belajar/selalu bikin onar') [PRIORITAS]"),
    101: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    102: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    103: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    104: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    105: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    106: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    107: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
    108: (0, 1, "bocor_indonesia", "LLM draft: Indonesia murni"),
}

assert len(J) == 108 and set(J) == set(range(1, 109)), "judgment dict must cover all 108"

HDR_OTENTIK = "OTENTIK? (1=ya,0=tidak)"
HDR_JELAS = "JELAS_HATE? (1=ya,0=tidak)"
HDR_MASALAH = "MASALAH"
HDR_CATATAN = "CATATAN"
HDR_NO = "no"

BANNER_FILL = PatternFill("solid", fgColor="FCE4D6")  # light orange = attention


def fill(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb["Validasi"]
    hdr = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(hdr)}
    c_no, c_ot, c_jh, c_ms, c_ct = (col[HDR_NO], col[HDR_OTENTIK], col[HDR_JELAS],
                                    col[HDR_MASALAH], col[HDR_CATATAN])

    # map row number -> `no` value
    no_to_row = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, c_no).value
        if v is not None and v != "":
            no_to_row[int(v)] = r

    filled = 0
    for no, (ot, jh, ms, ct) in J.items():
        r = no_to_row[no]
        ws.cell(r, c_ot).value = ot
        ws.cell(r, c_jh).value = jh
        ws.cell(r, c_ms).value = ms
        ws.cell(r, c_ct).value = ct
        # mark LLM-origin cells visibly (light orange tint on the two judgment cols)
        ws.cell(r, c_ot).fill = BANNER_FILL
        ws.cell(r, c_jh).fill = BANNER_FILL
        filled += 1

    # Warning banner on the Petunjuk sheet (first place a human looks)
    ins = wb["Petunjuk"]
    ins.insert_rows(1, amount=4)
    warnings = [
        "&&& PERHATIAN KHUSUS (2026-07-06) &&&&&&&&&&&&&&&&&&&&&&&&&&&&&&",
        "Kolom OTENTIK? / JELAS_HATE? / MASALAH / CATATAN sudah DIISI DRAFT oleh Claude (LLM, BUKAN penutur native),",
        "atas instruksi Bapak. Ini BUKAN validasi native. Semua nilai HANYA SARAN - mohon periksa & UBAH setiap baris",
        "sesuai penilaian Anda. Sel berwarna jingga = nilai draft LLM. Catatan LLM diawali 'LLM draft:' (boleh ditimpa).",
    ]
    for i, txt in enumerate(warnings, start=1):
        cell = ins.cell(i, 1, txt)
        cell.font = Font(bold=True)
        cell.fill = BANNER_FILL

    wb.save(path)
    print(f"{path.name}: filled {filled}/{len(J)} rows; Petunjuk banner added.")


for f in FILES:
    fill(f)
print("Done.")
