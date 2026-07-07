"""Build BLIND validation forms for a second/third native validator (E1, STRATEGY.md).

Why this exists: the paper's authenticity rates (97%/56%/11%, 55% overall) come from a
single validator (the first author, Mukhlis) who also designed the taxonomy and reviewed
the generation pipeline -- an n=1 result with no inter-rater reliability. To move past a
single-evaluator estimate we need 1-2 more native Javanese speakers (Yekti Asmoro Kanthi,
Daniel Rudiaman Sijabat -- confirmed active speakers, 2026-07-06) to independently judge
the SAME 108 items, BLIND to Mukhlis's existing verdicts.

Reads the same _key.csv that rebuild_form.py already produced (no need to regenerate
generation data). Produces one clean .xlsx per named validator with NO pre-filled
authenticity column (unlike VALIDATION_FORM.xlsx, which Mukhlis has already completed --
that file is untouched by this script).

Two judgment columns per row (not one), per STRATEGY.md Sec.6 point 1 -- authenticity and
hate-clarity are different questions and conflating them into a single "OTENTIK" column
makes it impossible to tell whether a detector "miss" is a real failure or a genuinely
ambiguous case:
  OTENTIK?    (1=ya,0=tidak)  -- same question/wording Mukhlis answered, for direct
                                 pairwise / Krippendorff-alpha comparability against his
                                 existing scored form.
  JELAS_HATE? (1=ya,0=tidak)  -- NEW, independent of authenticity: would a competent
                                 native speaker clearly read this as hate (attack on
                                 group identity), independent of how authentically
                                 Javanese it sounds? Lets us separate "not real Javanese"
                                 from "real Javanese but not actually hate" (cf. paper
                                 example #21, where sarcasm read as sincere praise).

Run: python experiments/generation_pilot/build_multivalidator_forms.py [names...]
Default names: yekti daniel
Writes: VALIDATION_FORM_<name>.xlsx for each name that doesn't already exist (never
overwrites a form that already has answers -- same safety rule as rebuild_form.py).
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

HERE = Path(__file__).parent
KEY = HERE / "_key.csv"

HEADERS = ["no", "model", "niche", "target", "register", "TEKS", "mekanisme",
           "machine_caught", "auto_concern", "PRIORITAS",
           "OTENTIK? (1=ya,0=tidak)", "JELAS_HATE? (1=ya,0=tidak)", "MASALAH", "CATATAN"]


def load_key() -> list[dict]:
    if not KEY.exists():
        raise SystemExit("Missing _key.csv. Run rebuild_form.py first (should already exist).")
    with KEY.open(encoding="utf-8") as f:
        return list(csv.DictReader(f))


def form_already_answered(path: Path) -> bool:
    if not path.exists():
        return False
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False
    wb = load_workbook(path)
    if "Validasi" not in wb.sheetnames:
        return False
    ws = wb["Validasi"]
    hdr = [c.value for c in ws[1]]
    if "OTENTIK? (1=ya,0=tidak)" not in hdr:
        return False
    gcol = hdr.index("OTENTIK? (1=ya,0=tidak)") + 1
    return any(ws.cell(r, gcol).value not in (None, "") for r in range(2, ws.max_row + 1))


def build_one(name: str, rows: list[dict]) -> None:
    out_path = HERE / f"VALIDATION_FORM_{name}.xlsx"
    if form_already_answered(out_path):
        print(f"SKIP {out_path.name}: already has answers filled in. "
              f"Score it first (score_multivalidator.py) or move it aside before rebuilding.")
        return

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Validasi"
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).fill = PatternFill("solid", fgColor="DDDDDD")
    prio_fill = PatternFill("solid", fgColor="FFF2CC")
    for r in rows:
        ws.append([r["no"], r["model"], r["niche"], r["intended_target"], r["register"],
                   r["text"], r["mekanisme"], r["machine_caught"], r["auto_concern"],
                   r["prioritas"] or "", "", "", "", ""])
        if r["prioritas"] == "1":
            for c in range(1, len(HEADERS) + 1):
                ws.cell(ws.max_row, c).fill = prio_fill
    widths = {"A": 5, "B": 10, "C": 20, "D": 16, "E": 9, "F": 58, "G": 44,
              "H": 9, "I": 22, "J": 9, "K": 14, "L": 14, "M": 16, "N": 26}
    for col, wd in widths.items():
        ws.column_dimensions[col].width = wd
    for rr in range(2, ws.max_row + 1):
        ws.cell(rr, 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(rr, 7).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    n_prio = sum(1 for r in rows if r["prioritas"] == "1")
    ins = wb.create_sheet("Petunjuk")
    guide = [
        [f"VALIDASI KEASLIAN — untuk {name.title()} (validator native ke-2/3, penutur Jawa aktif)"],
        [""],
        ["Ini bagian dari riset deteksi ujaran kebencian Jawa (tujuan DEFENSIF: menemukan celah "
         "detektor otomatis, bukan membuat ujaran kebencian). Semua 108 contoh sudah dinilai oleh "
         "Mukhlis (penulis pertama); form Anda TERPISAH dan BUTA terhadap jawabannya -- supaya kita "
         "bisa mengukur seberapa konsisten dua penutur asli menilai hal yang sama (inter-rater "
         "reliability). Estimasi waktu: 1-2 jam untuk 108 baris."],
        [""],
        [f"Total {len(rows)} contoh. Baris kuning (PRIORITAS=1, {n_prio} baris) = kalau waktu terbatas, isi INI dulu."],
        ["PRIORITAS = contoh yang lolos deteksi mesin (paling menarik utk temuan) atau di-flag panel-QC,"],
        ["plus 1 contoh per (model lokal x niche) untuk perbandingan antar-model."],
        [""],
        ["Untuk tiap baris isi 4 kolom kanan (K-N):"],
        ["K  OTENTIK?    -> 1 jika teks TERDENGAR seperti Bahasa Jawa asli pada register yang "
         "dimaksud (kolom 'register': ngoko/krama); 0 jika terasa aneh, textbook, atau bocor Indonesia."],
        ["L  JELAS_HATE? -> TERPISAH dari OTENTIK: 1 jika teks ini JELAS menyerang/merendahkan "
         "identitas kelompok (suku/agama/gender/politik) bagi penutur Jawa yang paham konteksnya; "
         "0 jika tidak jelas hate (mis. terbaca sebagai pujian tulus, sindiran ambigu, atau bukan "
         "menyerang kelompok). Sebuah teks BISA otentik (K=1) tapi TIDAK jelas hate (L=0), atau "
         "sebaliknya -- itu wajar, isi keduanya sesuai penilaian Anda masing-masing."],
        ["M  MASALAH     -> kalau OTENTIK=0, pilih salah satu: museum_krama | bocor_indonesia | "
         "bukan_hate | salah_register | salah_target | formulaic | lainnya"],
        ["N  CATATAN     -> opsional: alasan singkat, terutama kalau OTENTIK dan JELAS_HATE "
         "penilaiannya berbeda arah, atau kalau ragu."],
        [""],
        ["Kolom konteks (read-only, HANYA untuk bantu menilai -- JANGAN diubah, dan JANGAN dianggap "
         "sebagai jawaban benar):"],
        ["H  machine_caught = berapa dari 5 detektor otomatis menandai ini 'hate' (mis. 2/5). "
         "Kecil = lolos deteksi mesin."],
        ["I  auto_concern   = flag dari panel-QC LLM non-native (museum_krama/register_mismatch/"
         "formulaic/...) -- ini DUGAAN mesin, bukan mesti benar."],
        [""],
        ["Mohon isi TANPA mendiskusikan jawaban dengan Mukhlis atau validator lain dulu -- "
         "independensi penilaian adalah tujuan dari langkah ini."],
    ]
    for row in guide:
        ins.append(row)
    ins.column_dimensions["A"].width = 120
    for row_cells in ins.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    print(f"Wrote {out_path.name}: {len(rows)} rows, {n_prio} PRIORITAS.")


def main():
    names = sys.argv[1:] or ["yekti", "daniel"]
    rows = load_key()
    for name in names:
        build_one(name, rows)


if __name__ == "__main__":
    main()
